#!/usr/bin/env python3
"""Fast integrity checks for the public identity-remodelling release."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SIGNATURE_DIR = ROOT / "data" / "signatures"
FIGURE_DIR = ROOT / "figures" / "communications_biology_v2.1"
MAX_PUBLIC_FILE_BYTES = 50 * 1024 * 1024

MAIN_FIGURES = [
    "figure1_study_design_and_programme_derivation",
    "figure2_donor_disjoint_identity_remodelling",
    "figure3_external_recurrence_and_archival_transfer",
    "figure4_regulatory_and_epithelial_context",
    "figure5_genetic_perturbation_support",
    "figure6_reduced_measurement_candidate",
]
SUPP_FIGURES = [
    "figureS1_sampling_and_donor_stability",
    "figureS2_fine_state_and_composition_sensitivities",
    "figureS3_functional_and_regulatory_structure",
    "figureS4_external_and_ffpe_sensitivities",
    "figureS5_compact_derivation_and_benchmarks",
    "figureS6_multiomic_atlas_spatial_protein",
    "figureS7_empirical_perturbation_context",
]


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def close(observed: float, expected: float, tolerance: float = 1e-10) -> bool:
    return abs(observed - expected) <= tolerance


def check_definitions() -> None:
    ranking = pd.read_csv(SIGNATURE_DIR / "common_effect_ranking_8221.tsv.gz", sep="\t")
    programme = pd.read_csv(SIGNATURE_DIR / "state_aware_high_confidence_1843.tsv", sep="\t")
    candidates = pd.read_csv(SIGNATURE_DIR / "portable_candidates_53.tsv", sep="\t")
    reduced = pd.read_csv(SIGNATURE_DIR / "compact_candidate_8_genes.tsv", sep="\t")
    require(len(ranking) == 8221, "Expected an 8,221-gene common-effect ranking")
    require(len(programme) == 1843, "Expected 1,843 high-confidence genes")
    require(programme["shared_direction"].value_counts().to_dict() == {"down": 959, "up": 884}, "Unexpected programme arms")
    require(len(candidates) == 53, "Expected 53 measurable candidates")
    require(len(reduced) == 8, "Expected eight reduced-readout genes")
    require(reduced["arm"].value_counts().to_dict() == {"up": 4, "down": 4}, "Reduced candidate is not balanced")
    require(set(reduced["gene"]) == {"EPHB2", "REG1A", "LTBP1", "RNF43", "CALM2", "COX6C", "B2M", "ACAA2"}, "Reduced membership changed")
    require(set(reduced["gene"]).issubset(set(candidates["gene"])), "Reduced genes are outside the measurable universe")
    require(not reduced["validation_outcomes_used"].astype(bool).any(), "Validation leakage flag changed")
    require(reduced["panel_frozen_before_validation"].astype(bool).all(), "Freeze flag changed")


def check_principal_results() -> None:
    donor = pd.read_csv(ROOT / "results/state_shared_revision_v2/donor_site/donor_disjoint_replication_summary.tsv", sep="\t").iloc[0]
    require(int(donor["n_testable_strict_genes"]) == 1646, "Donor-disjoint testable count changed")
    require(close(donor["common_direction_match_fraction"], 0.986026731470231), "Direction agreement changed")
    require(close(donor["discovery_validation_effect_spearman"], 0.912185029960394), "Discovery-validation fidelity changed")

    decomposition = pd.read_csv(ROOT / "results/state_shared_revision_v2/fine_state_models/programme_composition_decomposition.tsv", sep="\t")
    primary = decomposition.loc[
        decomposition["partition"].eq("validation")
        & decomposition["k"].eq(4)
        & decomposition["scope"].eq("all")
        & decomposition["decomposition_type"].eq("fine_state")
    ].set_index("component")
    total = primary.loc["total", "estimate"]
    composition = primary.loc["composition", "estimate"]
    within = primary.loc["within", "estimate"]
    require(close(total, composition + within), "Primary decomposition does not close")
    require(close(within / total, 0.7911514513877733, 1e-9), "Primary within-state fraction changed")

    meta = pd.read_csv(ROOT / "results/state_shared_revision_v2/external_meta/random_effects_meta_summary.tsv", sep="\t")
    full = meta.loc[meta["signature_id"].eq("state_shared_1843") & meta["excluded_cohort"].eq("__NONE__")].iloc[0]
    require(close(full["pooled_standardized_effect"], 1.90288083576639), "External pooled effect changed")

    coverage = pd.read_csv(ROOT / "results/state_aware_program_v1/full_program_coverage_audit/full_programme_feature_coverage_summary.tsv", sep="\t")
    require(len(coverage) == 4, "Expected four full-programme coverage layers")
    require((coverage[["minimum_up_coverage", "minimum_down_coverage"]].min(axis=1) >= 0.75).all(), "An eligible projection arm fell below 75% coverage")

    becker = pd.read_csv(ROOT / "results/state_aware_program_v1/extended_validation_full_programme/becker/becker_full_programme_adjusted_models.tsv", sep="\t")
    becker_polyp = becker.loc[becker["outcome"].eq("epi__ca_route_signature") & becker["term"].eq("disease_stage_group_polyp")].iloc[0]
    require(close(becker_polyp["coef"], 0.6729875920516115), "Becker adjusted full-programme coefficient changed")

    atlas = pd.read_csv(ROOT / "results/state_aware_program_v1/extended_validation_full_programme/crc_atlas/atlas_locked_leave_one_study_out.tsv", sep="\t")
    atlas_polyp = atlas.loc[atlas["omitted_study"].eq("__NONE__") & atlas["state"].eq("polyp_epithelial")].iloc[0]
    require(close(atlas_polyp["coef"], 0.20995839710992545), "CRC Atlas adjusted full-programme coefficient changed")
    atlas_loo = atlas.loc[
        ~atlas["omitted_study"].eq("__NONE__")
        & atlas["outcome"].eq("score__ca_route_signature")
        & atlas["state"].eq("polyp_epithelial")
        & atlas["estimable"].astype(bool)
    ]
    require(len(atlas_loo) == 33 and (atlas_loo["coef"] > 0).all(), "CRC Atlas leave-one-study-out support changed")

    spatial = pd.read_csv(ROOT / "results/state_aware_program_v1/extended_validation_full_programme/perturbation_spatial/spatial_full_programme_tests.tsv", sep="\t")
    adjusted = spatial.loc[spatial["feature"].eq("route_residual_prolif_epithelial")].iloc[0]
    require(int(adjusted["n_sections"]) == 6 and int(adjusted["n_positive"]) == 6, "Adjusted spatial direction changed")
    require(close(adjusted["p_wilcoxon"], 0.03125), "Adjusted spatial test changed")

    apc = pd.read_csv(ROOT / "results/perturbation_validation_locked_route/gse125472_contrast_summary.tsv", sep="\t")
    require(not apc.empty, "APC-knockout donor contrasts are missing")

    internal = pd.read_csv(ROOT / "results/state_shared_revision_v2/compact_rank/random_eight_gene_benchmark_summary.tsv", sep="\t").iloc[0]
    external = pd.read_csv(ROOT / "results/state_shared_revision_v2/external_rank/random_eight_gene_benchmark_summary.tsv", sep="\t").iloc[0]
    require(internal["observed_compact_spearman"] > internal["random_q95"], "Internal reduced-readout benchmark boundary changed")
    require(external["observed_median_cohort_spearman"] < external["random_q95"], "External non-uniqueness boundary changed")


def check_figures_and_workbooks() -> None:
    for stem in MAIN_FIGURES + SUPP_FIGURES:
        for suffix in ("pdf", "png", "svg"):
            require((FIGURE_DIR / f"{stem}.{suffix}").is_file(), f"Missing figure: {stem}.{suffix}")
    source = load_workbook(ROOT / "data/source_data/Source_Data_v2.1.xlsx", read_only=True)
    supplement = load_workbook(ROOT / "data/supplementary/Supplementary_Tables_1-11_v2.1.xlsx", read_only=True)
    require(len(source.sheetnames) == 77, "Unexpected Source Data worksheet count")
    require(len(supplement.sheetnames) == 70, "Unexpected supplementary worksheet count")
    require("README" in source.sheetnames and "SOURCE_MAP" in source.sheetnames, "Source Data index sheets are missing")
    require("README" in supplement.sheetnames, "Supplementary workbook README is missing")
    require(all(any(name.startswith(f"T{number}") for name in supplement.sheetnames) for number in range(1, 12)), "A supplementary table group is missing")
    forbidden = re.compile(r"historical|legacy|genki|virtual|287|12_gene", flags=re.IGNORECASE)
    require(not any(forbidden.search(name) for name in source.sheetnames + supplement.sheetnames), "A retired analysis entered a current workbook")
    source.close()
    supplement.close()


def check_required_files_and_hygiene() -> None:
    required = [
        "README.md",
        "CITATION.cff",
        "environment.yml",
        "data/datasets.tsv",
        "workflow/pipeline.tsv",
        "analysis/state_aware_build_discovery_pseudobulk_v1.R",
        "analysis/state_aware_integrate_common_effects_v1.R",
        "analysis/state_shared_revision_define_fine_states_v2.py",
        "analysis/state_shared_revision_fine_state_models_v2.R",
        "analysis/audit_state_shared_full_program_coverage_v1.py",
        "analysis/validate_state_shared_full_programme_extended_layers_v2.py",
        "analysis/contracts/state_shared_full_programme_projection_addendum_v1_2026-08-30.md",
        "analysis/derive_state_shared_compact_panel_v1.R",
        "analysis/build_state_shared_revision_figure6_v3.R",
    ]
    missing = [path for path in required if not (ROOT / path).is_file()]
    require(not missing, "Missing required files: " + ", ".join(missing))
    oversized = [str(path.relative_to(ROOT)) for path in ROOT.rglob("*") if path.is_file() and ".git" not in path.parts and path.stat().st_size > MAX_PUBLIC_FILE_BYTES]
    require(not oversized, "Files exceed 50 MiB: " + ", ".join(oversized))
    local_home = "/home/" + "elliottlv"
    forbidden_text = re.compile(re.escape(local_home) + r"|ghp_[A-Za-z0-9]{30,}|github_pat_[A-Za-z0-9_]{30,}")
    hits: list[str] = []
    for path in ROOT.rglob("*"):
        if not path.is_file() or ".git" in path.parts or path.suffix.lower() not in {".md", ".py", ".r", ".json", ".txt", ".tsv", ".yml", ".yaml", ".cff"}:
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        if forbidden_text.search(text):
            hits.append(str(path.relative_to(ROOT)))
    require(not hits, "Machine path or credential-like text found: " + ", ".join(hits))
    governed_paths = [
        ROOT / "data/source_data/figureS2g_dslab_common_composition.tsv",
        FIGURE_DIR / "source_data/figureS2g_dslab_common_composition.tsv",
    ]
    require(not any(path.exists() for path in governed_paths), "Governed DSLab source file is public")


def main() -> None:
    checks = [
        ("frozen definitions", check_definitions),
        ("principal numerical results", check_principal_results),
        ("figures and workbooks", check_figures_and_workbooks),
        ("required files and hygiene", check_required_files_and_hygiene),
    ]
    for label, function in checks:
        function()
        print(f"[PASS] {label}")
    print(f"Release verification passed: {len(checks)}/{len(checks)} checks")


if __name__ == "__main__":
    main()
